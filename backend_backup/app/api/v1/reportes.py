from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import Optional
from datetime import date, datetime
import io
from openpyxl import Workbook
from fpdf import FPDF
from app.database import get_db
from app.models.expediente import Expediente
from app.models.cliente import Cliente
from app.models.contrato import Contrato
from app.models.transaccion import Transaccion

router = APIRouter()
TIPOS = ["expedientes", "clientes", "contratos", "finanzas"]

async def get_data(db, tipo, desde=None, hasta=None):
    if tipo == "expedientes":
        q = select(Expediente).order_by(Expediente.id)
        if desde: q = q.where(Expediente.fecha_apertura >= desde)
        if hasta: q = q.where(Expediente.fecha_apertura <= hasta)
        return (await db.execute(q)).scalars().all()
    elif tipo == "clientes":
        return (await db.execute(select(Cliente).order_by(Cliente.id))).scalars().all()
    elif tipo == "contratos":
        q = select(Contrato).order_by(Contrato.id)
        if desde: q = q.where(Contrato.fecha_inicio >= desde)
        if hasta: q = q.where(Contrato.fecha_inicio <= hasta)
        return (await db.execute(q)).scalars().all()
    elif tipo == "finanzas":
        q = select(Transaccion).order_by(Transaccion.fecha.desc())
        if desde: q = q.where(Transaccion.fecha >= datetime(desde.year, desde.month, desde.day))
        if hasta: q = q.where(Transaccion.fecha <= datetime(hasta.year, hasta.month, hasta.day, 23, 59, 59))
        return (await db.execute(q)).scalars().all()
    raise HTTPException(400, "Tipo inválido")

def generate_excel(data, tipo):
    wb = Workbook()
    ws = wb.active
    ws.title = tipo.capitalize()
    if tipo == "expedientes":
        ws.append(["ID", "Número", "Cliente", "Juzgado", "Estado", "Prioridad", "Apertura"])
        for item in data:
            ws.append([item.id, item.numero_expediente, item.cliente, item.juzgado or "", item.estado, item.prioridad, str(item.fecha_apertura)])
    elif tipo == "clientes":
        ws.append(["ID", "Nombre", "Email", "Teléfono", "RFC"])
        for item in data:
            ws.append([item.id, item.nombre, item.email, item.telefono or "", item.rfc or ""])
    elif tipo == "contratos":
        ws.append(["ID", "Número", "Cliente ID", "Inicio", "Vencimiento", "Estado"])
        for item in data:
            ws.append([item.id, item.numero_contrato, item.cliente_id, str(item.fecha_inicio), str(item.fecha_vencimiento), item.estado])
    elif tipo == "finanzas":
        ws.append(["ID", "Tipo", "Categoría", "Monto", "Fecha"])
        for item in data:
            ws.append([item.id, item.tipo, item.categoria, item.monto, str(item.fecha)])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf(data, tipo):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)
    if tipo == "expedientes":
        for item in data:
            pdf.cell(0, 10, f"{item.numero_expediente} - {item.cliente} ({item.estado})", new_x="LMARGIN", new_y="NEXT")
    elif tipo == "clientes":
        for item in data:
            pdf.cell(0, 10, f"{item.nombre} - {item.email}", new_x="LMARGIN", new_y="NEXT")
    elif tipo == "contratos":
        for item in data:
            pdf.cell(0, 10, f"{item.numero_contrato} - {item.estado}", new_x="LMARGIN", new_y="NEXT")
    elif tipo == "finanzas":
        for item in data:
            pdf.cell(0, 10, f"{item.tipo}: ${item.monto:,.2f}", new_x="LMARGIN", new_y="NEXT")
    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return output

@router.get("/excel")
async def reporte_excel(tipo: str = Query(...), desde: Optional[date] = None, hasta: Optional[date] = None, db: AsyncSession = Depends(get_db)):
    if tipo not in TIPOS: raise HTTPException(400, "Tipo inválido")
    data = await get_data(db, tipo, desde, hasta)
    excel_bytes = generate_excel(data, tipo)
    return StreamingResponse(excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=reporte_{tipo}.xlsx"})

@router.get("/pdf")
async def reporte_pdf(tipo: str = Query(...), desde: Optional[date] = None, hasta: Optional[date] = None, db: AsyncSession = Depends(get_db)):
    if tipo not in TIPOS: raise HTTPException(400, "Tipo inválido")
    data = await get_data(db, tipo, desde, hasta)
    pdf_bytes = generate_pdf(data, tipo)
    return StreamingResponse(pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=reporte_{tipo}.pdf"})
